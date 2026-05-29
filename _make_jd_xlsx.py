"""Generate Excel summary for Jaedaeng branch GPS coordinates."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRANCHES = [
    # นุชจรินทร์ (พี่กัส)
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (ธนิยะ)",                      13.7301775, 100.5307144),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (ดุสิต เซ็นทรัล พาร์ค)",       13.7283064, 100.534965),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (พรานนก)",                     13.7535468, 100.4455994),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (Paseo กาญจนาภิเษก)",          13.7665982, 100.4036471),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (เซ็นทรัลพระราม 2)",           13.662599,  100.4347654),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (The Bright พระราม 2)",        13.6687652, 100.4497944),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (ศาลายา ปั๊ม PT)",              13.7887022, 100.3393726),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (ปั๊ม ปตท.เวสวิลเลจ)",         13.8567282, 100.407688),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (ปตท.บรมราชชนนี)",             13.7819421, 100.4105383),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (บางจาก ราชพฤกษ์)",            13.7907227, 100.4446842),
    ("นุชจรินทร์", "พี่กัส",   "ร้านส้มตำเจ๊แดง (PTT ราชพฤกษ์)",                13.8408768, 100.4077997),
    # คณัสวรรณ (พี่เฟิร์น)
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (มอเตอร์เวย์-ชลบุรี)",         13.5393791, 101.006768),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (เอกเกณฑ์ พัทยา)",             12.9510719, 100.8857879),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (The Street รัชดา)",           13.7535468, 100.4455941),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (PTT. นวลจันทร์)",             13.8182153, 100.6462443),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (ปตท. พระราม 4)",              13.7131324, 100.5818959),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (ต้นซุง)",                     None, None),
    ("คณัสวรรณ",   "พี่เฟิร์น", "ร้านส้มตำเจ๊แดง (ฟอร์จุนทาวน์)",               13.759541,  100.5649858),
    # ขวัญดาว (พี่ดาว)
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (สยาม)",                        13.7353013, 100.5108902),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (สยามพารากอน)",                13.7463,    100.5346),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (เซ็นทรัลเวิลด์)",             13.7447218, 100.4645232),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (คู่บอน)",                     13.8548134, 100.67797),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (บางจากสุขุมวิท 62)",          13.6952912, 100.6033721),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (ทากะทาวน์)",                  13.7453243, 100.5700709),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (Emporium)",                   13.7428975, 100.564992),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (True Digital Park)",          13.6866139, 100.6348432),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (For you park บางนา)",         13.6712974, 100.6183834),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (เดอะมอลล์ งามวงศ์วาน)",       13.8558865, 100.5390694),
    ("ขวัญดาว",    "พี่ดาว",   "ร้านส้มตำเจ๊แดง (เซ็นทรัลอีสต์วิลล์)",         13.8029718, 100.6143198),
    # พิมพ์พิชชา (พี่อ้อย)
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (Charn at the Avenue)",         13.8007595, 100.3735592),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (ปตท. สายไหม 56)",              13.9184717, 100.6576505),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (บางจากรังสิต คลองสอง)",       13.9885509, 100.6373587),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (ฟิวเจอร์ปาร์ค รังสิต)",        13.9892012, 100.6148237),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (เมืองทองธานี)",                13.9147566, 100.5371567),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (เทพารักษ์-สายไหม)",            13.8812783, 100.6239798),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (ปั๊มบางจาก-ศาลายา)",          13.8068012, 100.3129815),
    ("พิมพ์พิชชา", "พี่อ้อย", "ร้านส้มตำเจ๊แดง (ประชาชื่น)",                   13.8417627, 100.5307696),
    ("พิมพ์พิชชา", "พี่อ้อย", "4007 ซีคอนศรีนครินทร์ (เจ๊แดง จุ่มนัว)",        13.6932864, 100.6453704),
    ("พิมพ์พิชชา", "พี่อ้อย", "4008 เซ็นทรัลศาลายา (เจ๊แดง จุ่มนัว)",          13.7873021, 100.2733108),
    ("พิมพ์พิชชา", "พี่อ้อย", "4018 เทอมินอล21 พระราม 3 (เจ๊แดง จุ่มนัว)",     13.689139,  100.4295476),
    # จุรีพร (พี่อีฟ)
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (S-Oasis)",                     None, None),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (ท็อปส์ เซ็นทรัลลาดพร้าว)",     13.8159219, 100.5583448),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (Little Walk รัตนาธิเบศร์)",    13.8650669, 100.4921928),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (The Circle ราชพฤกษ์)",        13.7672438, 100.4397739),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (ไอคอนสยาม)",                  13.7264086, 100.500041),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (กรุงเทพกริฑา)",                13.7537365, 100.7145711),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (ปั๊มบางจาก รามคำแหง)",        13.7789767, 100.6750197),
    ("จุรีพร",     "พี่อีฟ",  "ร้านส้มตำเจ๊แดง (เสนานิคม)",                    13.8341619, 100.5799618),
    ("จุรีพร",     "พี่อีฟ",  "4001 แฟชั่น ไอส์แลนด์ (เจ๊แดง จุ่มนัว)",       13.8253442, 100.6788238),
    ("จุรีพร",     "พี่อีฟ",  "4005 ซีคอนบางแค (เจ๊แดง จุ่มนัว)",              13.7123589, 100.4308969),
    ("จุรีพร",     "พี่อีฟ",  "4015 เดอะมอลล์ท่าพระ (เจ๊แดง จุ่มนัว)",         13.7139355, 100.477666),
]

wb = Workbook()

# ===== Sheet 1: ทุกสาขา =====
ws = wb.active
ws.title = "สาขาทั้งหมด"

thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
header_font = Font(name="Sarabun", size=12, bold=True, color="FFFFFF")
cell_font = Font(name="Sarabun", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["#", "BZM (ชื่อจริง)", "BZM (นิคเนม)", "ชื่อสาขา", "Latitude", "Longitude", "สถานะพิกัด", "Google Maps"]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

missing_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ok_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

for i, (mgr, nick, name, lat, lng) in enumerate(BRANCHES, 1):
    has_coord = lat is not None and lng is not None
    status = "✅ มีพิกัด" if has_coord else "❌ ไม่มี"
    gmaps = f'=HYPERLINK("https://maps.google.com/?q={lat},{lng}","เปิดแผนที่")' if has_coord else ""
    row = [i, mgr, nick, name, lat, lng, status, gmaps]
    ws.append(row)
    r = i + 1
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c)
        cell.font = cell_font
        cell.border = border
        if c == 1 or c == 7:
            cell.alignment = center
        else:
            cell.alignment = left
        if not has_coord:
            cell.fill = missing_fill

# Column widths
widths = [5, 16, 12, 50, 13, 13, 14, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 28

# ===== Sheet 2: สรุป =====
ws2 = wb.create_sheet("สรุปตาม BZM")
ws2.append(["BZM (ชื่อจริง)", "BZM (นิคเนม)", "จำนวนสาขา", "มีพิกัด", "ขาดพิกัด"])
for col_idx in range(1, 6):
    cell = ws2.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# Group by BZM
from collections import defaultdict
groups = defaultdict(lambda: {"total": 0, "has": 0, "missing": 0, "mgr": "", "nick": ""})
for mgr, nick, name, lat, lng in BRANCHES:
    g = groups[nick]
    g["mgr"] = mgr
    g["nick"] = nick
    g["total"] += 1
    if lat is not None and lng is not None:
        g["has"] += 1
    else:
        g["missing"] += 1

for nick, g in groups.items():
    ws2.append([g["mgr"], g["nick"], g["total"], g["has"], g["missing"]])

total_row = len(groups) + 2
ws2.append(["รวมทั้งหมด", "", len(BRANCHES), sum(g["has"] for g in groups.values()), sum(g["missing"] for g in groups.values())])
for c in range(1, 6):
    cell = ws2.cell(row=total_row, column=c)
    cell.font = Font(name="Sarabun", size=11, bold=True)
    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    cell.border = border
    cell.alignment = center if c >= 3 else left

for r in range(2, total_row):
    for c in range(1, 6):
        cell = ws2.cell(row=r, column=c)
        cell.font = cell_font
        cell.border = border
        cell.alignment = center if c >= 3 else left

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.row_dimensions[1].height = 28

# ===== Sheet 3: สาขาที่ขาดพิกัด =====
ws3 = wb.create_sheet("ขาดพิกัด GPS")
ws3.append(["#", "BZM (ชื่อจริง)", "BZM (นิคเนม)", "ชื่อสาขา", "หมายเหตุ"])
for col_idx in range(1, 6):
    cell = ws3.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

n = 0
for mgr, nick, name, lat, lng in BRANCHES:
    if lat is None or lng is None:
        n += 1
        ws3.append([n, mgr, nick, name, "ต้องเพิ่มพิกัดผ่าน Hub: 🍜 จัดการสาขาเจ๊แดง"])
        for c in range(1, 6):
            cell = ws3.cell(row=n + 1, column=c)
            cell.font = cell_font
            cell.border = border
            cell.alignment = center if c in (1,) else left
            cell.fill = missing_fill

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 50
ws3.column_dimensions["E"].width = 50
ws3.row_dimensions[1].height = 28

out_path = "JAEDAENG-LOCATIONS.xlsx"
wb.save(out_path)
print(f"OK saved: {out_path}")
print(f"Total: {len(BRANCHES)} · มีพิกัด {sum(1 for b in BRANCHES if b[3] is not None)} · ขาด {sum(1 for b in BRANCHES if b[3] is None)}")
